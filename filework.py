import sys
from os import path
from openpyxl import Workbook, load_workbook
from random import randint

PASSWORD = "p4SSw0rd"


try:
	WORDS = [x.strip() for x in open("data/bip39.txt").readlines()]
	METAMASK = "data/MetaMask-Chrome.crx"
	DRIVER = "data/chromedriver-win64/chromedriver.exe"
except FileNotFoundError:
	bundle_dir = getattr(sys, '_MEIPASS', path.abspath(path.dirname(__file__)))
	WORDS = [x.strip() for x in open(path.abspath(path.join(bundle_dir, "data/bip39.txt"))).readlines()]
	METAMASK = path.abspath(path.join(bundle_dir, "data/MetaMask-Chrome.crx"))
	DRIVER = path.abspath(path.join(bundle_dir, "data/chromedriver-win64/chromedriver.exe"))

try:
	WB = load_workbook('found.xlsx')
	WS = WB.active
except FileNotFoundError:
	WB = Workbook()
	WS = WB.active
	WS.append(["Address", "First trans", "Balance", "SRP"])


def gen_phrase(length: int = 12) -> list:
	return [WORDS[randint(0, 2047)] for _ in range(length)]


def saver(wallet: dict):
	WS.append([wallet["addr"], wallet["age"], wallet["bal"], wallet["srp"]])
	WB.save("found.xlsx")
